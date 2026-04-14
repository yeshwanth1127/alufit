import uuid
import csv
from zipfile import BadZipFile
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.orm import Session

from app.models.entities import BoqLineItem, BoqVersion, BoqVersionStatus


def import_boq_from_xlsx(db: Session, boq_version: BoqVersion, file_bytes: bytes) -> tuple[int, list[str]]:
    if boq_version.status != BoqVersionStatus.draft:
        raise ValueError("BOQ version is locked")
    db.query(BoqLineItem).filter(BoqLineItem.boq_version_id == boq_version.id).delete()
    errors: list[str] = []
    rows_iter: Any
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(min_row=2, values_only=True)
    except (BadZipFile, InvalidFileException, OSError):
        # Fallback for non-xlsx files (csv/tsv/txt). Header row is still treated as first row.
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError as e:
            raise ValueError(
                "Unsupported file format. Upload an .xlsx file or a UTF-8 CSV/TSV text file."
            ) from e
        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            raise ValueError("BOQ file is empty")
        sample = "\n".join(lines[:10])
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            class _FallbackDialect(csv.excel):
                delimiter = ","

            dialect = _FallbackDialect
        reader = csv.reader(lines, dialect)
        next(reader, None)  # Skip header row for parity with Excel flow.
        rows_iter = reader
    sort_order = 0
    count = 0
    for row in rows_iter:
        if not row or all(v is None or str(v).strip() == "" for v in row[:4]):
            continue
        line_no = str(row[0]).strip() if row[0] is not None else ""
        desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        uom = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
        try:
            qty = float(row[3]) if len(row) > 3 and row[3] is not None else 0.0
        except (TypeError, ValueError):
            errors.append(f"Row {sort_order + 2}: invalid quantity")
            qty = 0.0
        try:
            rate = float(row[4]) if len(row) > 4 and row[4] is not None else 0.0
        except (TypeError, ValueError):
            errors.append(f"Row {sort_order + 2}: invalid rate")
            rate = 0.0
        amount = qty * rate
        if len(row) > 5 and row[5] is not None:
            try:
                amount = float(row[5])
            except (TypeError, ValueError):
                pass
        if not line_no or not desc:
            errors.append(f"Row {sort_order + 2}: missing line_no or description")
            continue
        item = BoqLineItem(
            boq_version_id=boq_version.id,
            line_no=line_no,
            description=desc,
            uom=uom or None,
            quantity=qty,
            rate=rate,
            amount=amount,
            sort_order=sort_order,
        )
        db.add(item)
        sort_order += 1
        count += 1
    boq_version.row_count_snapshot = count
    db.flush()
    return count, errors


def lock_boq_version(db: Session, boq_version: BoqVersion, actor_id: uuid.UUID | None) -> BoqVersion:
    from datetime import datetime, timezone

    if boq_version.status == BoqVersionStatus.locked:
        return boq_version
    boq_version.status = BoqVersionStatus.locked
    boq_version.locked_at = datetime.now(timezone.utc)
    db.flush()
    db.refresh(boq_version)
    return boq_version
